from pathlib import Path
import re

import openpyxl as xl
import pytest
from openpyxl.workbook.defined_name import DefinedName
from typer.testing import CliRunner

from named_xlsx.cli import app

ANSI_ESCAPE_RE = re.compile(r"\x1b\[[0-9;]*m")


def _plain(text: str) -> str:
    return ANSI_ESCAPE_RE.sub("", text)


@pytest.fixture()
def workbook_path(tmp_path: Path) -> Path:
    path = tmp_path / "cli-typer.xlsx"
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["A2"] = 2
    wb.defined_names.add(DefinedName("i.a", attr_text="Sheet1!$A$1"))
    wb.defined_names.add(DefinedName("i.b", attr_text="Sheet1!$A$2"))
    wb.save(path)
    wb.close()
    return path


def test_root_help_lists_commands():
    runner = CliRunner()
    result = runner.invoke(app, ["--help"])
    assert result.exit_code == 0
    assert "save" in result.stdout
    assert "load" in result.stdout
    assert "spec" in result.stdout
    assert "refresh" in result.stdout


def test_save_subcommand_supports_hyphenated_options(workbook_path: Path):
    runner = CliRunner()
    result = runner.invoke(app, ["save", str(workbook_path), "--filter-prefix", "i."])
    assert result.exit_code == 0
    assert '"i.a" = 1' in result.stdout
    assert '"i.b" = 2' in result.stdout


def test_save_subcommand_supports_legacy_underscore_options(
    workbook_path: Path, tmp_path: Path
):
    runner = CliRunner()
    out_path = tmp_path / "saved.toml"
    result = runner.invoke(
        app,
        ["save", str(workbook_path), "--filter_prefix", "i.", "--p_out", str(out_path)],
    )
    assert result.exit_code == 0
    assert out_path.read_text().startswith('[Sheet1]\n"i.a" = 1')


def test_spec_subcommand_help_mentions_hyphenated_option():
    runner = CliRunner()
    result = runner.invoke(app, ["spec", "--help"])
    assert result.exit_code == 0
    assert "--filter-prefix" in _plain(result.stdout)


def test_load_subcommand_rejects_read_only_engine(workbook_path: Path, tmp_path: Path):
    runner = CliRunner()
    cfg_path = tmp_path / "update.toml"
    out_path = tmp_path / "out.xlsx"
    cfg_path.write_text('[Sheet1]\n"i.a" = 10\n')
    result = runner.invoke(
        app,
        [
            "load",
            str(cfg_path),
            str(workbook_path),
            str(out_path),
            "--engine",
            "Calamine",
        ],
    )
    assert result.exit_code != 0
    assert "read-only" in result.output
