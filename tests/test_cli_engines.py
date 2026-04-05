from pathlib import Path

import openpyxl as xl
import pytest
from openpyxl.workbook.defined_name import DefinedName

from named_xlsx.cli import load, save, specifications


@pytest.fixture()
def workbook_path(tmp_path: Path) -> Path:
    path = tmp_path / "engines.xlsx"
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["A2"] = 2
    wb.defined_names.add(DefinedName("i.a", attr_text="Sheet1!$A$1"))
    wb.defined_names.add(DefinedName("i.b", attr_text="Sheet1!$A$2"))
    wb.save(path)
    wb.close()
    return path


def test_save_accepts_engine_name(workbook_path: Path, capsys):
    save(workbook_path, engine="OpenPYXL")
    out = capsys.readouterr().out
    assert '"i.a" = 1' in out
    assert '"i.b" = 2' in out


def test_specifications_accepts_calamine_engine_name(workbook_path: Path, capsys):
    specifications(workbook_path, engine="Calamine")
    out = capsys.readouterr().out
    assert "i.a" in out
    assert "Sheet1" in out


def test_load_rejects_read_only_engine(workbook_path: Path, tmp_path: Path):
    cfg_path = tmp_path / "update.toml"
    out_path = tmp_path / "out.xlsx"
    cfg_path.write_text('[Sheet1]\n"i.a" = 10\n')
    with pytest.raises(ValueError, match="read-only"):
        load(cfg_path, workbook_path, out_path, engine="Calamine")


def test_unknown_engine_name_is_rejected(workbook_path: Path):
    with pytest.raises(ValueError, match="Unknown engine"):
        save(workbook_path, engine="MissingEngine")
