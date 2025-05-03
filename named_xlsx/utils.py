# coding=utf-8
from dataclasses import dataclass, field
from itertools import product
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from parse import parse
from memoization import cached
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils.cell import cols_from_range, rows_from_range, get_column_letter


@dataclass
class Table:
    name: str
    sheet: str
    _table: openpyxl.worksheet.table.Table = field(repr=False)

    @property
    def cells(self):
        return self._table.ref

    def mapper_columns(self, **kwargs) -> dict[str, tuple[str, str]]:
        table_colnames = self._table.column_names
        cr = CellRange(self.cells)
        cr.shrink(**kwargs)
        table_colsindx = [col for row, col in cr.top]
        table_rowsindx = [row for row, col in cr.left]
        table_row_top, table_row_bottom = table_rowsindx[0], table_rowsindx[-1]

        data = {}
        for col_name, col_indx in zip(table_colnames, table_colsindx):
            col_letter = get_column_letter(col_indx)
            addr = f"{col_letter}{table_row_top}:{col_letter}{table_row_bottom}"
            data[col_name] = (self.sheet, addr)
        return data


@cached
def table_destination(
    ref: str, /, tables: dict[str, Table], **kwargs
) -> tuple[str, str]:
    r = parse("{table_name}[{table_colname}]", ref)
    table_name = r["table_name"]
    table_colname = r["table_colname"]
    return tables[table_name].mapper_columns(**kwargs)[table_colname]


@cached
def get_tables(wb) -> dict[str, Table]:
    tables = {}
    for sheet_name in wb.sheetnames:
        for table_name in wb[sheet_name].tables:
            table = wb[sheet_name].tables[table_name]
            tables[table_name] = Table(
                name=table_name,
                sheet=sheet_name,
                _table=table,
            )
    return tables


@cached
def get_destinations(
    defined_name: DefinedName, tables: dict[str, Table]
) -> list[tuple[str, str]]:
    """
    Get plain spreadsheet address from defined names.

    This is a middle-man function because openpyxl does not seem to deal with
    defined names that point to table columns.
    As per project conventions, tables are labelled `t.<name>`,
    so this function deals with that case separately with a bunch of
    hard-set conditions.

    """
    if defined_name.attr_text.startswith("t."):
        # Has a table as per project convention: with header and total rows
        # Get destination, shrinking range to data cells (without header and total rows).
        dest = defined_name.attr_text
        return [table_destination(dest, tables=tables, top=1, bottom=1)]
    return list(defined_name.destinations)


@dataclass(frozen=True)
class XLSXAddress:
    """
    Excel address, with or without sheet name, according to project conventions.

    >>> a = XLSXAddress("My Sheet!A10:A15")
    >>> a
    XLSXAddress(value='My Sheet!A10:A15')
    >>> a.sheet
    'My Sheet'
    >>> a.coord
    'A10:A15'
    >>> a = XLSXAddress("A10:A15")
    >>> a
    XLSXAddress(value='A10:A15')
    >>> a.sheet
    >>> a.coord
    'A10:A15'

    """

    value: str
    sheet: str = field(init=False, repr=False)
    coord: str = field(init=False, repr=False)
    _parsed: CellRange = field(init=False, repr=False)

    def __post_init__(self):

        value = self.value
        if " " in self.value and "!" in self.value:
            sheet, coord = self.value.rsplit("!")
            value = f"'{sheet}'!{coord}"
        try:
            _parsed = CellRange(value)
        except ValueError as e:
            raise ValueError(f"Cannot parse: {self.value=}") from e
        object.__setattr__(self, "_parsed", _parsed)
        object.__setattr__(self, "sheet", self._parsed.title)
        object.__setattr__(self, "coord", self._parsed.coord)

    @classmethod
    def from_parts(cls, sheet: str, coords: str):
        """
        Build instance from parts.

        Parameters
        ----------
        sheet
            Sheet name. Can be `None`.
        coords
            Coordinates. Must always be given.

        Examples
        --------
        >>> XLSXAddress.from_parts("My Sheet", "A10")
        XLSXAddress(value='My Sheet!A10')
        >>> XLSXAddress.from_parts(None, "A10")
        XLSXAddress(value='A10')

        """
        return cls(coords if sheet is None else f"{sheet}!{coords}")

    def as_array(self, order: str = "row", squeeze: bool = True) -> np.ndarray:
        """
        Generate coordinate array (row major: left-right then top-down).

        Parameters
        ----------
        order
            Specify the order, one of ['row', 'col'].
        squeeze
            Drop single-dimensional entries from the shape of the resulting array.

        Examples
        --------
        >>> XLSXAddress("A10:B11").as_array()
        array([['A10', 'B10'],
               ['A11', 'B11']], dtype='<U3')
        >>> XLSXAddress("Hello!A10:B11").as_array()
        array([['A10', 'B10'],
               ['A11', 'B11']], dtype='<U3')

        """
        d = dict(row=rows_from_range, col=cols_from_range)
        out = np.array(list(d[order](self.coord)))
        if squeeze:
            out = np.squeeze(out)
        return out

    @property
    def is_range(self) -> bool:
        """

        Examples
        --------
        >>> XLSXAddress("A10").is_range
        False
        >>> XLSXAddress("A10:B11").is_range
        True

        """
        if self.size == 1:
            return False
        if self.size > 1:
            return True
        raise ValueError(f"Unexpected {self.size=} for {self.value=}")

    @property
    def size(self) -> int:
        """
        Size of the address (number of elements).

        Examples
        --------
        >>> XLSXAddress("A10").size
        np.int64(1)
        >>> XLSXAddress("A10:A10").size
        np.int64(1)
        >>> XLSXAddress("A10:C11").size
        np.int64(6)
        >>> XLSXAddress("A10:A12").size
        np.int64(3)

        """
        return np.prod(list(self._parsed.size.values()))

    @property
    def shape(self) -> tuple[int, int]:
        """
        Shape of the address as `(row, column)`.

        Examples
        --------
        >>> XLSXAddress("A10").shape
        (1, 1)
        >>> XLSXAddress("A10:A10").shape
        (1, 1)
        >>> XLSXAddress("A10:C11").shape
        (2, 3)
        >>> XLSXAddress("A10:A12").shape
        (3, 1)

        """
        d = self._parsed.size
        return d["rows"], d["columns"]

    def format(self):
        return f"{self.sheet}!{self.coord}" if self.sheet is not None else self.coord


def compare_sheets(
    path1,
    sheet1: str,
    path2,
    sheet2: str | None = None,
    start_row: int = 0,
    start_col: int = 0,
):
    if sheet2 is None:
        sheet2 = sheet1
    df1 = pd.read_excel(path1, sheet_name=sheet1, header=None)
    df2 = pd.read_excel(path2, sheet_name=sheet2, header=None)
    if df1.shape == df2.shape:
        shape = df1.shape
    else:
        print(
            f"Shape mismatch: {df1.shape=} != {df2.shape}. Reducing to overlapping area"
        )
        shape = min(df1.shape[0], df2.shape[0]), min(df1.shape[1], df2.shape[1])
    arr1 = df1.to_numpy()
    arr2 = df2.to_numpy()
    for irow, icol in product(range(start_row, shape[0]), range(start_col, shape[1])):
        val1 = arr1[irow, icol]
        val2 = arr2[irow, icol]
        report = False
        if pd.isnull(val1):
            if pd.isnull(val2):
                continue
            report = True
        elif isinstance(val1, str):
            report = val1 != str(val2)
        elif isinstance(val1, (int, float)):
            try:
                report = not np.isclose(val1, val2)
            except Exception:
                report = True
        if report:
            ecol = get_column_letter(icol + 1)
            erow = irow + 1
            print(f"{sheet1}!{ecol}{erow}: {val1} != {val2}")


def nanaverage(arr: np.ndarray, weights: np.ndarray | None = None) -> float:
    """
    Compute average ignoring `nan` values in `arr` and `weights`.

    Examples
    --------
    >>> nanaverage(np.array([1, 2, 3]))
    2.0
    >>> nanaverage(np.array([np.nan, 2, 3]))
    2.5
    >>> nanaverage(np.array([np.nan, 2, 3]), np.array([1, 2, np.nan]))
    2.0

    """
    msg = "Argument `{}` must be a numpy array (got `{}`)."
    if not isinstance(arr, np.ndarray):
        raise ValueError(msg.format("arr", type(arr)))
    if weights is not None and not isinstance(weights, np.ndarray):
        raise ValueError(msg.format("weights", type(weights)))
    if weights is None:
        weights = np.full_like(arr, fill_value=1)
    indices = ~np.isnan(arr) & ~np.isnan(weights)
    return np.average(arr[indices], weights=weights[indices]).item()


Numeric = int | float | np.ndarray | pd.Series
Pathlike = str | Path
Addresslike = str | XLSXAddress
